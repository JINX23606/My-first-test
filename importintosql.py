from openpyxl import load_workbook
import mysql.connector
#excel
workbook = load_workbook('import.xlsx')
sheet = workbook.active

values = []
for row in sheet.iter_rows(min_row=2,values_only=True):
    print(row)
    values.append(row)

#database
db = mysql.connector.connect(
    host ='localhost',
    port = 3306,
    user = 'root',
    password = 'Bellic4825550100#',
    database = 'ink_want_to_buy'
)

cursor = db.cursor()
sql = '''
    INSERT INTO products(title,price,is_necessary)
    VALUE (%s,%s,%s);
'''
cursor.executemany(sql,values)
db.commit()
print('เพิ่มข้อมูลจำนวน '+str(cursor.rowcount)+' แถว')